from openpyxl import *
import openpyxl

server_names = ["mqqspadc", "appdadc"]
realsit = []
for names in server_names:
    # opening server report
    name = names+".xlsx"
    wb = openpyxl.load_workbook(name)
    sheetnames = wb.sheetnames
    sheetnames.remove("Sudo")
    wssudo = wb["Sudo"]
    # creating sudo users dictonary
    sudodict = {}
    sudolen = len(wssudo["A"])
    for i in range(1, sudolen+1):
        col1 = "A"+str(i)
        col2 = "E"+str(i)
        sudodict[wssudo[col1].value] = wssudo[col2].value
    # opening applicaiton users list
    for sheet in sheetnames:
        wsapp = wb[sheet]
        applen = len(wsapp["A"])
        templist = []
        # appending data to lists

        for i in range(2, applen+1):
            cola = "A"+str(i)
            colb = "B"+str(i)
            colc = "C"+str(i)
            cold = "D"+str(i)
            templist.append(wsapp[cola].value)
            templist.append(wsapp[colb].value)
            templist.append(wsapp[colc].value)
            templist.append(wsapp[cold].value)
            if str(wsapp[cola].value) in sudodict:
                templist.append("Yes")
                templist.append(sudodict[wsapp[cola].value])
            else:
                templist.append("NO")
            templist.insert(0, names)
            templist.insert(0, "")
            realsit.append(templist)
            print(realsit)
            templist = []
        wb1 = openpyxl.load_workbook("final.xlsx")
        ws1 = wb1[sheet]
        for row in realsit:
            ws1.append(row)
        wb1.save("final.xlsx")
        realsit = []
    wb.save(name)

wb2 = openpyxl.load_workbook("final.xlsx")
sheets = wb2.sheetnames
i = 2
for sh in sheetnames:
    sn = 1
    ws2 = wb2[sh]
    wslen = len(ws2["A"])
    for k in range(i, wslen+1):
        col = "A"+str(k)
        ws2[col] = sn
        sn = sn+1
wb2.save("final.xlsx")
